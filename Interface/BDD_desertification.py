import os
import pandas as pd
import sqlite3
import openpyxl

repertoire_script = os.path.dirname(os.path.abspath(__file__))
chemin1= os.path.abspath(os.path.join(repertoire_script,".."))

#chemin fichier excell
chemin_excel_1= chemin1+r"/detection_lac1/Etude_de_la_flotte.xlsx"
chemin_excel_2= chemin1+r"/detection_lac1/marque_de_flotte.xlsx"
chemin_excel_3= chemin1+r"/detection_lac1/Surfaces_au_complet.xlsx"

chemin_excel=[
            chemin1+r"/detection_lac1/Etude_de_la_flotte.xlsx",
            chemin1+r"/detection_lac1/marque_de_flotte.xlsx",
            chemin1+r"/detection_lac1/Surfaces_au_complet.xlsx"
            ]

# Nom de la base de données SQLite à créer
nom_db_1 = 'Etude_de_la_flotte.db'
nom_db_2 = 'marque_de_flotte.db'
nom_db_3 = 'Surfaces_au_complet.db'

nom_db =[
        repertoire_script + '/Etude_de_la_flotte.db',
        repertoire_script + '/marque_de_flotte.db',
        repertoire_script + '/Surfaces_au_complet.db'
        ]

#Nom de la table
nom_tb_1 ='etude'
nom_tb_2 ='marque'
nom_tb_3 ='surface'

nom_tb=[
        'etude',
        'marque',
        'surface'
        ]

def implémenter_id_1(nom_db, chemin_excel):



    # Lire le fichier Excel
    wb = openpyxl.load_workbook(chemin_excel)
    sheet = wb.active

    conn = sqlite3.connect(nom_db)
    cursor = conn.cursor()

    # Vérifier si la table existe déjà
    cursor.execute("SELECT count(name) FROM sqlite_master WHERE type='table' AND name='etude'")
    table_exists = cursor.fetchone()[0]

    if table_exists == 0:
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS DATA")
        cursor.execute("""CREATE TABLE etude(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            Image TEXT,
            Proportion_de_blanc INTEGER,
            Surface_en_km² INTEGER)
            """)
    
    # Parcourir chaque ligne du fichier Excel
    for row in sheet.iter_rows(values_only=True):
        # Insérer les données dans la base de données
        cursor.execute("INSERT INTO etude (Image , Proportion_de_blanc, Surface_en_km²) VALUES (?, ?, ?)", row)
    
    # Valider la transaction et fermer la connexion
    conn.commit()
    conn.close()

def implémenter_id_2(nom_db, chemin_excel):

    # Lire le fichier Excel
    wb = openpyxl.load_workbook(chemin_excel)
    sheet = wb.active

    conn = sqlite3.connect(nom_db)
    cursor = conn.cursor()

    # Vérifier si la table existe déjà
    cursor.execute("SELECT count(name) FROM sqlite_master WHERE type='table' AND name='marque'")
    table_exists = cursor.fetchone()[0]

    if table_exists == 0:
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS DATA")
        cursor.execute("""CREATE TABLE marque(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            Nom TEXT,
            Nb_kmcube_au_km2 INTEGER,
            Hauteur_moy_arbre_m INTEGER,
            Prix_au_litre INTEGER)
            """)
    # Parcourir chaque ligne du fichier Excel
    for row in sheet.iter_rows(values_only=True):
        # Insérer les données dans la base de données
        cursor.execute("INSERT INTO marque (Nom , Nb_kmcube_au_km2, Hauteur_moy_arbre_m, Prix_au_litre) VALUES (?, ?, ?, ?)", row)
    
    # Valider la transaction et fermer la connexion
    conn.commit()
    conn.close()

def implémenter_id_3(nom_db, chemin_excel):

    # Lire le fichier Excel
    wb = openpyxl.load_workbook(chemin_excel)
    sheet = wb.active

    conn = sqlite3.connect(nom_db)
    cursor = conn.cursor()

    # Vérifier si la table existe déjà
    cursor.execute("SELECT count(name) FROM sqlite_master WHERE type='table' AND name='surface'")
    table_exists = cursor.fetchone()[0]

    if table_exists == 0:
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS DATA")
        cursor.execute("""CREATE TABLE surface(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            Nom_de_l_arbre TEXT,
            Noms_associés INTEGER,
            Image TEXT,
            Résultat INTEGER,
            Résultat_Hauteur INTEGER,
            Nb_arbre_total INTEGER)
            """)

    # Parcourir chaque ligne du fichier Excel
    for row in sheet.iter_rows(values_only=True):
        # Insérer les données dans la base de données
        cursor.execute("INSERT INTO surface (Nom_de_l_arbre, Noms_associés, Image, Résultat, Résultat_Hauteur, Nb_arbre_total) VALUES (?, ?, ?, ?, ?, ?)", row)
    
    # Valider la transaction et fermer la connexion
    conn.commit()
    conn.close()

i = 0 
for i in range (3) :
    print(i,"top")
    if i == 0:
        implémenter_id_1(nom_db[i],chemin_excel[i])
    if i == 1:
        implémenter_id_2(nom_db[i],chemin_excel[i])
    if i == 2:
        implémenter_id_3(nom_db[i],chemin_excel[i])
    else:
        print("finish")
    i = i+1
    print(i)