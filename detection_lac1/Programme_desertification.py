"""
Cet algorithme décrit de manière détaillée les étapes du programme principal. Voici une explication de chaque étape :

1. **Créer un classeur Excel (Workbook) et une feuille active (Worksheet)** : Cette étape initialise un classeur Excel et une feuille de calcul pour enregistrer les résultats.

2. **Appeler la fonction `process_images_in_folder`** : Cette fonction est appelée pour traiter les images dans un dossier spécifié. Les images sont traitées en appliquant différentes opérations telles que le seuillage pour détecter les pixels verts.

3. **Parcourir les sous-dossiers et rechercher les fichiers "medium_canny.png"** : Cette étape parcourt les sous-dossiers à la recherche de fichiers spécifiques. Pour chaque fichier trouvé, les contours sont détectés et les informations de superficie sont calculées.

4. **Sauvegarder le classeur Excel** : Une fois toutes les données collectées, le classeur Excel est sauvegardé avec un nom spécifié.

5. **Charger les données des fichiers Excel dans des DataFrames** : Les données extraites des fichiers Excel sont chargées dans des structures de données appelées DataFrames, généralement utilisées dans la bibliothèque Pandas pour manipuler et analyser des données tabulaires.

6. **Créer un DataFrame pour stocker les résultats finaux** : Un DataFrame est créé pour stocker les résultats finaux obtenus à partir des calculs effectués.

7. **Calculer les résultats pour chaque ligne d'arbres et de surfaces** : Les informations sur les arbres et les surfaces sont utilisées pour calculer plusieurs mesures, telles que le coût par kilomètre carré, la hauteur totale et le nombre total d'arbres.

8. **Enregistrer les résultats dans un nouveau fichier Excel** : Les résultats sont enregistrés dans un nouveau fichier Excel avec un nom spécifique.

9. **Créer une interface tkinter pour afficher les résultats par essence d'arbres** : Une interface utilisateur est créée à l'aide du module tkinter pour afficher les résultats filtrés par essence d'arbres.

10. **Afficher les résultats lorsque l'utilisateur sélectionne une essence et clique sur le bouton** : Lorsque l'utilisateur sélectionne une essence d'arbres et clique sur un bouton, les résultats correspondants sont affichés dans l'interface utilisateur.

11. **Enregistrer les chemins des images dans un fichier "chemins_images.txt"** : Les chemins des images sont enregistrés dans un fichier texte pour référence ultérieure.

12. **Modifier les chemins pour les passer en relatif** : Les chemins absolus des images sont modifiés pour les rendre relatifs par rapport à un dossier de base.

13. **Modifier un fichier HTML en utilisant la fonction `modifier_fichier`** : Un fichier HTML est modifié en ajoutant des balises HTML `<img>` autour des chemins des images.

14. **Ouvrir le fichier HTML dans le navigateur par défaut** : Le fichier HTML modifié est ouvert dans le navigateur par défaut pour affichage.

Cet algorithme décrit de manière séquentielle les actions que le programme principal effectue pour traiter les images, calculer les résultats et afficher les informations à l'utilisateur.
"""

import os
import cv2
import sqlite3
import openpyxl
import webbrowser
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

####################################################################################################################
def modifier_fichier(input_path, output_path):
    """
    Lit un fichier d'entrée, ajoute des balises HTML <img> autour de chaque ligne,
    puis écrit le résultat dans un fichier de sortie.

    Args:
        input_path (str): Le chemin du fichier d'entrée.
        output_path (str): Le chemin du fichier de sortie.
    """
    with open(input_path, 'r') as fichier_entree:
        lignes = fichier_entree.readlines()

    lignes_modifiees = ['<img src="..\{}">'.format(ligne.strip()) for ligne in lignes]

    with open(output_path, 'w') as fichier_sortie:
        fichier_sortie.write('\n'.join(lignes_modifiees))

# Doctest pour la fonction modifier_fichier
def test_modifier_fichier():
    """
    >>> input_path_test = 'test_input.txt'
    >>> output_path_test = 'test_output.txt'
    >>> with open(input_path_test, 'w') as f:
    ...     f.write('image1.jpg\\nimage2.png\\nimage3.gif')
    >>> modifier_fichier(input_path_test, output_path_test)
    >>> with open(output_path_test, 'r') as f:
    ...     modified_content = f.read()
    >>> print(modified_content)
    <img src="image1.jpg">
    <img src="image2.png">
    <img src="image3.gif">
    >>> os.remove(input_path_test)
    >>> os.remove(output_path_test)
    """
    pass

####################################################################################################################
def creer_table_1():
    conn = sqlite3.connect('désertification.db')
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS DATA")
    cur.execute("""CREATE TABLE DESERTIFICATION(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chemin TEXT,
        nom INTEGER,
        surface INTEGER,
        hauteur INTEGER,
        volume INTEGER)
        """)
    conn.close


def trouver_chemins_images(dossier, extensions=('jpg', 'jpeg', 'png', 'gif')):
    """
    Recherche récursivement tous les fichiers d'images dans un dossier.

    Args:
        dossier (str): Le chemin du dossier à parcourir.
        extensions (tuple): Les extensions de fichiers à rechercher. Par défaut, ('jpg', 'jpeg', 'png', 'gif').

    Returns:
        List[str]: Une liste contenant les chemins complets de tous les fichiers d'images trouvés.
    """
    chemins_images = []
    for dossier_actuel, sous_dossiers, fichiers in os.walk(dossier):
        for fichier in fichiers:
            if fichier.lower().endswith(extensions):
                chemin_complet = os.path.join(dossier_actuel, fichier)
                chemins_images.append(chemin_complet)
    return chemins_images

# Doctest pour la fonction trouver_chemins_images
def test_trouver_chemins_images():
    """
    >>> import tempfile
    >>> tmpdir = tempfile.mkdtemp()
    >>> img1_path = os.path.join(tmpdir, 'image1.jpg')
    >>> img2_path = os.path.join(tmpdir, 'image2.png')
    >>> img3_path = os.path.join(tmpdir, 'subfolder', 'image3.gif')
    >>> os.makedirs(os.path.join(tmpdir, 'subfolder'))
    >>> open(img1_path, 'w').close()
    >>> open(img2_path, 'w').close()
    >>> open(img3_path, 'w').close()

    >>> chemins = trouver_chemins_images(tmpdir)
    >>> len(chemins)
    3
    >>> img1_path in chemins
    True
    >>> img2_path in chemins
    True
    >>> img3_path in chemins
    True

    >>> img4_path = os.path.join(tmpdir, 'document.txt')
    >>> open(img4_path, 'w').close()
    >>> chemins = trouver_chemins_images(tmpdir)
    >>> len(chemins)
    3
    >>> img4_path in chemins
    False

    >>> img5_path = os.path.join(tmpdir, 'image5.jpg')
    >>> os.makedirs(os.path.join(tmpdir, 'empty_folder'))
    >>> open(img5_path, 'w').close()
    >>> chemins = trouver_chemins_images(os.path.join(tmpdir, 'empty_folder'))
    >>> len(chemins)
    0

    >>> import shutil
    >>> shutil.rmtree(tmpdir)
    """
    pass

####################################################################################################################

def enregistrer_chemins_dans_fichier(chemins_images, fichier_sortie='chemins_images.txt'):
    """
    Enregistre les chemins des images dans un fichier texte.

    Args:
        chemins_images (List[str]): Une liste contenant les chemins complets des fichiers d'images.
        fichier_sortie (str): Le chemin du fichier de sortie. Par défaut, 'chemins_images.txt'.
    """
    with open(fichier_sortie, 'w') as fichier:
        for chemin in chemins_images:
            fichier.write(chemin + '\n')

# Doctest pour la fonction enregistrer_chemins_dans_fichier
def test_enregistrer_chemins_dans_fichier():
    """
    >>> import tempfile
    >>> tmpdir = tempfile.mkdtemp()
    >>> img1_path = os.path.join(tmpdir, 'image1.jpg')
    >>> img2_path = os.path.join(tmpdir, 'image2.png')
    >>> img3_path = os.path.join(tmpdir, 'subfolder', 'image3.gif')
    >>> os.makedirs(os.path.join(tmpdir, 'subfolder'))
    >>> open(img1_path, 'w').close()
    >>> open(img2_path, 'w').close()
    >>> open(img3_path, 'w').close()

    >>> chemins = [img1_path, img2_path, img3_path]
    >>> fichier_sortie = os.path.join(tmpdir, 'output.txt')
    >>> enregistrer_chemins_dans_fichier(chemins, fichier_sortie)

    >>> with open(fichier_sortie, 'r') as f:
    ...     contenu_fichier = f.read()
    >>> print(contenu_fichier)
    {0}
    {1}
    {2}

    >>> shutil.rmtree(tmpdir)
    """
    pass

####################################################################################################################


def modifier_chemins_dans_fichier(input_path, output_path, dossier_base):
    """
    Modifie les chemins absolus dans un fichier pour les rendre relatifs par rapport à un dossier de base.

    Args:
        input_path (str): Chemin du fichier d'entrée contenant les chemins absolus.
        output_path (str): Chemin du fichier de sortie pour enregistrer les chemins relatifs.
        dossier_base (str): Dossier de base pour rendre les chemins relatifs.
    """
    with open(input_path, 'r') as fichier_entree:
        lignes = fichier_entree.readlines()

    chemins_relatifs = [os.path.relpath(chemin.strip(), dossier_base) for chemin in lignes]

    with open(output_path, 'w') as fichier_sortie:
        fichier_sortie.write('\n'.join(chemins_relatifs))


# Doctest pour la fonction modifier_chemins_dans_fichier
def test_modifier_chemins_dans_fichier():
    """
    >>> import tempfile
    >>> tmpdir = tempfile.mkdtemp()
    >>> img1_path = os.path.join(tmpdir, 'image1.jpg')
    >>> img2_path = os.path.join(tmpdir, 'image2.png')
    >>> img3_path = os.path.join(tmpdir, 'subfolder', 'image3.gif')
    >>> os.makedirs(os.path.join(tmpdir, 'subfolder'))
    >>> open(img1_path, 'w').close()
    >>> open(img2_path, 'w').close()
    >>> open(img3_path, 'w').close()

    >>> chemins_absolus = [img1_path, img2_path, img3_path]
    >>> fichier_entrée = os.path.join(tmpdir, 'input.txt')
    >>> fichier_sortie = os.path.join(tmpdir, 'output.txt')
    
    >>> with open(fichier_entrée, 'w') as f:
    ...     f.write('\\n'.join(chemins_absolus))

    >>> dossier_base = tmpdir
    >>> modifier_chemins_dans_fichier(fichier_entrée, fichier_sortie, dossier_base)

    >>> with open(fichier_sortie, 'r') as f:
    ...     chemins_relatifs = f.read().splitlines()
    
    >>> print(chemins_relatifs)
    ['image1.jpg', 'image2.png', 'subfolder/image3.gif']

    >>> shutil.rmtree(tmpdir)
    """
    pass



####################################################################################################################

def ouvrir_html(chemin_fichier):
    """
    Ouvre le fichier HTML dans le navigateur par défaut.

    Args:
        chemin_fichier (str): Le chemin complet du fichier HTML.

    Example:
        >>> import tempfile
        >>> import webbrowser
        >>> tmpdir = tempfile.mkdtemp()
        >>> fichier_html = os.path.join(tmpdir, 'test.html')
        >>> open(fichier_html, 'w').close()
        >>> ouvrir_html(fichier_html)
        >>> # Note: La fonction webbrowser.open ne renvoie pas de valeur directe,
        >>> # mais elle devrait ouvrir le fichier dans le navigateur par défaut.

        >>> shutil.rmtree(tmpdir)
    """
    webbrowser.open('file://' + chemin_fichier)


####################################################################################################################

def process_images_in_folder(input_folder, output_folder, lower_green, upper_green):
    """
    Traite les images dans un dossier, applique la détection de couleur verte,
    crée des images en niveaux de gris, applique un flou gaussien,
    effectue une détection de contour Canny à différents seuils, et enregistre les résultats.

    Args:
        input_folder (str): Le chemin du dossier d'entrée contenant les images.
        output_folder (str): Le chemin du dossier de sortie pour enregistrer les résultats.
        lower_green (tuple): Les composantes HSV inférieures de la plage de vert à détecter.
        upper_green (tuple): Les composantes HSV supérieures de la plage de vert à détecter.
    """
    # Créer les dossiers de sortie s'ils n'existent pas
    os.makedirs(output_folder, exist_ok=True)
    
    # Obtenir la liste de tous les fichiers image dans le dossier d'entrée
    image_files = [f for f in os.listdir(input_folder) if f.endswith(('.jpg', '.png', '.jpeg', '.bmp', '.PNG'))]

    for image_file in image_files:
        # Charger l'image
        image_path = os.path.join(input_folder, image_file)
        image = cv2.imread(image_path)
        if image is None:
            print(f"Échec de la lecture de {image_file}. Ignorer...")
            continue

        print(f"Traitement de {image_file}...")

        # Convertir l'image en format HSV (teinte, saturation, valeur)
        hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

        # Définir la plage de couleurs verte que vous souhaitez détecter
        lower_green = np.array(lower_green)
        upper_green = np.array(upper_green)

        # Créer un masque pour les pixels verts dans la plage spécifiée
        mask = cv2.inRange(hsv_image, lower_green, upper_green)

        # Appliquer le masque à l'image d'origine
        result = cv2.bitwise_and(image, image, mask=mask)

        # Convertir en niveaux de gris
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Enregistrer l'image en niveaux de gris
        output_path = os.path.join(output_folder, image_file)
        os.makedirs(output_path, exist_ok=True)
        cv2.imwrite(os.path.join(output_path, "gray_image.png"), gray)

        # Appliquer un flou gaussien
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)

        # Appliquer la détection de contours Canny avec différents seuils
        large = cv2.Canny(blurred, 10, 200)
        medium = cv2.Canny(blurred, 30, 150)
        narrow = cv2.Canny(blurred, 240, 250)

        # Enregistrer les images traitées dans le dossier de sortie
        cv2.imwrite(os.path.join(output_path, "large_canny.png"), large)
        cv2.imwrite(os.path.join(output_path, "medium_canny.png"), medium)
        cv2.imwrite(os.path.join(output_path, "narrow_canny.png"), narrow)
        cv2.imwrite(os.path.join(output_path, "Green_and_Black.png"), result)
        image = cv2.imread(os.path.join(output_path, "Green_and_Black.png"))
        trouver_contour(output_path,image)
        print(f"{image_file} traité et les résultats ont été enregistrés dans {output_path}")

def trouver_contour(output_path,image):
    # Convertir l'image en espace de couleur HSV
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Définir la plage de couleur bleue dans l'espace de couleur HSV
    lower_green = np.array([35, 50, 50])  # Plage inférieure (vert)
    upper_green = np.array([85, 255, 255]) # Plage supérieure (vert)

    # Binariser l'image en utilisant une plage de couleur bleue
    mask = cv2.inRange(hsv, lower_green, upper_green)

    # Trouver les contours dans l'image binarisée
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # Si des contours sont trouvés
    if contours:
        # Trouver le plus grand contour
        biggest_contour = max(contours, key=cv2.contourArea)

        # Tracer le contour du plus grand contour sur l'image originale
        cv2.drawContours(image, [biggest_contour], -1, (255, 0, 0), 2)

        # Enregistrer l'image avec le contour tracé
        cv2.imwrite(os.path.join(output_path, "Blue_with_contour.png"), image)

    # Si des contours sont trouvés
    if contours:
        # Trouver le plus grand contour
        biggest_contour = max(contours, key=cv2.contourArea)

        # Créer un masque noir de la même taille que l'image originale
        mask_filled = np.zeros_like(image)

        # Remplir la zone contourée avec une couleur spécifique (bleu dans cet exemple)
        cv2.fillPoly(mask_filled, [biggest_contour], (255, 0, 0))

        # Superposer l'image originale avec le masque rempli
        result = cv2.addWeighted(image, 1, mask_filled, 0.5, 0)

        # Enregistrer l'image avec le contour rempli
        cv2.imwrite(os.path.join(output_path, "Green_with_filled_contour.png"), result)


# Doctest pour la fonction process_images_in_folder
def test_process_images_in_folder():
    """
    >>> input_folder_test = 'test_input_images'
    >>> output_folder_test = 'test_output_images'
    >>> lower_green_test = (40, 40, 40)
    >>> upper_green_test = (80, 255, 255)
    >>> os.makedirs(input_folder_test, exist_ok=True)
    >>> os.makedirs(output_folder_test, exist_ok=True)
    >>> image_path_test = os.path.join(input_folder_test, 'test_image.jpg')
    >>> cv2.imwrite(image_path_test, np.zeros((100, 100, 3), dtype=np.uint8))  # Crée une image noire
    >>> process_images_in_folder(input_folder_test, output_folder_test, lower_green_test, upper_green_test)
    >>> # Vérifier si les fichiers résultants ont été créés dans le dossier de sortie
    >>> assert os.path.isfile(os.path.join(output_folder_test, 'test_image', 'gray_image.png'))
    >>> assert os.path.isfile(os.path.join(output_folder_test, 'test_image', 'large_canny.png'))
    >>> assert os.path.isfile(os.path.join(output_folder_test, 'test_image', 'medium_canny.png'))
    >>> assert os.path.isfile(os.path.join(output_folder_test, 'test_image', 'narrow_canny.png'))
    >>> assert os.path.isfile(os.path.join(output_folder_test, 'test_image', 'Green_and_Black.png'))
    >>> # Nettoyer les fichiers de test
    >>> os.remove(image_path_test)
    >>> os.rmdir(os.path.join(output_folder_test, 'test_image'))
    >>> os.rmdir(output_folder_test)
    >>> os.rmdir(input_folder_test)
    """
    pass

####################################################################################################################

def detecter_contours_et_remplir(image_path):
    """
    Détecte les contours blancs dans une image, remplit les zones encloses,
    et enregistre l'image résultante.

    Args:
        image_path (str): Le chemin de l'image à traiter.

    Returns:
        np.ndarray or None: L'image remplie en blanc ou None si la lecture de l'image échoue.
    """
    # Charger l'image
    image = cv2.imread(image_path)
    if image is None:
        return None

    # Conversion en niveaux de gris
    image_gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Détection des contours blancs
    _, thresh = cv2.threshold(image_gray, 200, 255, cv2.THRESH_BINARY)

    # Dilatation
    kernel = np.ones((3, 3), np.uint8)
    thresh = cv2.dilate(thresh, kernel, iterations=5)

    # Érosion pour fermer les zones (moins agressivement)
    thresh = cv2.erode(thresh, kernel, iterations=5)

    # Trouver les contours
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # Créer une image vide pour le remplissage
    filled_image = np.zeros_like(image)

    # Remplir les contours en blanc
    cv2.drawContours(filled_image, contours, -1, (255, 255, 255), thickness=cv2.FILLED)

    # Enregistrement de l'image modifiée
    output_path = image_path.replace('medium_canny.png', 'filled.png')
    cv2.imwrite(output_path, filled_image)

    return filled_image

# Doctest pour la fonction detecter_contours_et_remplir
def test_detecter_contours_et_remplir():
    """
    >>> input_image_path = 'test_image_medium_canny.png'
    >>> output_image_path = 'test_image_filled.png'
    >>> # Créer une image avec un contour blanc
    >>> image = np.zeros((100, 100, 3), dtype=np.uint8)
    >>> cv2.rectangle(image, (10, 10), (90, 90), (255, 255, 255), thickness=2)
    >>> cv2.imwrite(input_image_path, image)
    >>> filled_image = detecter_contours_et_remplir(input_image_path)
    >>> assert filled_image is not None
    >>> assert np.all(filled_image[15, 15] == [255, 255, 255])  # Vérifier le remplissage à l'intérieur du contour
    >>> os.remove(input_image_path)
    >>> os.remove(output_image_path)
    """
    pass

####################################################################################################################

def compter_pixels_et_calculer_superficie_blanc(image):
    """
    Compte le nombre total de pixels, le nombre de pixels blancs et calcule la superficie en km²
    pour une image donnée.

    Args:
        image (numpy.ndarray): L'image à analyser, représentée comme un tableau NumPy.

    Returns:
        tuple or None: Un tuple contenant le nombre total de pixels, le nombre de pixels blancs
        et la superficie en km², ou None si l'image est None.
    """
    if image is None:
        return None

    # Compter le nombre total de pixels
    total_pixels = image.size

    # Compter le nombre de pixels blancs
    white_pixels = np.count_nonzero(np.all(image == [255, 255, 255], axis=2))

    # Calculer la superficie en km²
    superficie_km2 = (white_pixels / (38 * 38)) * 400_000_000

    return total_pixels, white_pixels, superficie_km2




def compter_pixels_bleus(image_path):
    # Charger l'image
    image = cv2.imread(image_path)

    # Convertir l'image en espace de couleur HSV
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Définir la plage de couleur bleue dans l'espace de couleur HSV
    lower_blue = np.array([90, 50, 50])
    upper_blue = np.array([130, 255, 255])

    # Binariser l'image en utilisant une plage de couleur bleue
    mask = cv2.inRange(hsv, lower_blue, upper_blue)

    # Compter le nombre de pixels bleus
    nombre_pixels_bleus = np.count_nonzero(mask)

    return nombre_pixels_bleus

# Doctest pour la fonction compter_pixels_et_calculer_superficie_blanc
def test_compter_pixels_et_calculer_superficie_blanc():
    """
    >>> import numpy as np
    >>> # Créer une image de 100x100 pixels avec un carré blanc au centre
    >>> image = np.zeros((100, 100, 3), dtype=np.uint8)
    >>> cv2.rectangle(image, (30, 30), (70, 70), (255, 255, 255), thickness=cv2.FILLED)
    >>> total_pixels, white_pixels, superficie_km2 = compter_pixels_et_calculer_superficie_blanc(image)
    >>> assert total_pixels == 30000  # 100x100x3
    >>> assert white_pixels == 1600   # (70-30)x(70-30)
    >>> assert superficie_km2 == 4210526.315789474
    """
    pass
####################################################################################################################
def supprimer_extension(nom_fichier):
    """
    Supprime l'extension d'un nom de fichier.

    Args:
        nom_fichier (str): Le nom du fichier avec ou sans extension.

    Returns:
        str: Le nom du fichier sans extension, ou le même nom si aucune extension n'est présente.
    """
    if '.' in nom_fichier:
        return nom_fichier.split('.')[0]
    else:
        return nom_fichier

# Doctest pour la fonction supprimer_extension
def test_supprimer_extension():
    """
    >>> supprimer_extension('fichier.txt')
    'fichier'
    >>> supprimer_extension('image.png')
    'image'
    >>> supprimer_extension('document')
    'document'
    >>> supprimer_extension('archive.tar.gz')
    'archive'
    >>> supprimer_extension('sans_extension')
    'sans_extension'
    """
    pass
####################################################################################################################

def main(nom_fichier):
    """
    Modifie un fichier Excel en remplaçant le contenu de la colonne 1 par les noms de fichiers sans extension.

    Args:
        nom_fichier (str): Le chemin du fichier Excel à modifier.

    Raises:
        Exception: Une exception est levée en cas d'erreur lors du traitement du fichier Excel.

    Example:
        Pour utiliser cette fonction, appelez-la avec le chemin du fichier Excel :
        >>> main('chemin/vers/votre/fichier.xlsx')
        Le fichier Excel a été modifié avec succès.
    """
    try:
        # Ouvrir le fichier Excel
        workbook = openpyxl.load_workbook(nom_fichier)
        sheet = workbook.active

        # Remplacer la cellule A1 par "Image"
        sheet['A1'] = "Image"

        # Parcourir les cellules de la colonne 1 et supprimer l'extension du contenu
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                contenu_cellule = cell.value
                if contenu_cellule:
                    cell.value = supprimer_extension(contenu_cellule)

        # Sauvegarder le fichier Excel modifié
        workbook.save(nom_fichier)
        print("Le fichier Excel a été modifié avec succès.")

    except Exception as e:
        print(f"Une erreur s'est produite : {str(e)}")

# Doctest pour la fonction main
def test_main():
    """
    >>> import openpyxl
    >>> import os
    >>> file_path = 'test_file.xlsx'
    >>> workbook = openpyxl.Workbook()
    >>> sheet = workbook.active
    >>> sheet['A1'] = 'test_file.jpg'
    >>> workbook.save(file_path)
    >>> main(file_path)
    Le fichier Excel a été modifié avec succès.
    >>> loaded_workbook = openpyxl.load_workbook(file_path)
    >>> loaded_sheet = loaded_workbook.active
    >>> cell_value = loaded_sheet['A1'].value
    >>> assert cell_value == 'test_file'
    >>> os.remove(file_path)
    """
    pass
####################################################################################################################

# Obtenez le répertoire actuel du script
repertoire_script = os.path.dirname(os.path.abspath(__file__))
# Remplacez 'chemin_du_dossier' par le chemin du dossier que vous souhaitez parcourir
dossier_a_parcourir = os.path.join(repertoire_script, 'seuillage')

# Remplacez 'chemins_images.txt' par le chemin complet du fichier de sortie si nécessaire
fichier_sortie = os.path.join(repertoire_script, 'interface', 'chemin.txt')

# Spécifiez le chemin complet de votre fichier HTML
chemin_fichier_html = os.path.join(repertoire_script, 'interface', 'testsupp.html')

# Répertoire racine contenant les sous-dossiers d'images
repertoire_racine = os.path.join(repertoire_script, 'seuillage')

# Création d'un classeur Excel
wb = Workbook()
ws = wb.active
# Ajout des entêtes
ws.append(["Nom de l'image", "Proportion de blanc (%)", "Surface en km²"])


if __name__ == "__main__":
    #creer_table_1()
    input_folder = os.path.join(repertoire_script, 'photo')
    output_folder = os.path.join(repertoire_script, 'seuillage')
    
    # Plage de couleurs verte en format HSV (teinte, saturation, valeur)
    lower_green = [35, 50, 50]  # Plage inférieure (vert)
    upper_green = [85, 255, 255]  # Plage supérieure (vert)
    
    process_images_in_folder(input_folder, output_folder,lower_green,upper_green)

    # Parcourez les sous-dossiers et recherchez "medium_canny.png"
    for dossier, sous_dossiers, fichiers in os.walk(repertoire_racine):
        for fichier in fichiers:
            if fichier == 'medium_canny.png':
                chemin_image = os.path.join(dossier, fichier)
                filled_image = detecter_contours_et_remplir(chemin_image)
                if filled_image is not None:
                    total_pixels, white_pixels, superficie_km2 = compter_pixels_et_calculer_superficie_blanc(filled_image)
                    print(f"Image traitée : {chemin_image}")
                    print(f"Nombre total de pixels : {total_pixels}")
                    print(f"Nombre de pixels blancs : {white_pixels}")
                    print(f"Superficie en km² : {superficie_km2} km²")
                    # Utilisez le nom du dossier au lieu du nom de l'image
                    #ws.append([os.path.basename(os.path.dirname(chemin_image)), (white_pixels / total_pixels) * 100, superficie_km2])


    
    reference_image_path = os.path.join(repertoire_script+r'\seuillage\04-2020.jpg\Green_with_filled_contour.png')
    superficie_reference_km2 = 340  # Superficie en km² de la zone bleue de l'image de référence
    pixel_bleu_reference = compter_pixels_bleus(reference_image_path)
    print(pixel_bleu_reference)

    # Parcourez les sous-dossiers et recherchez "Green_with_filled_contour.png"
    for dossier, sous_dossiers, fichiers in os.walk(repertoire_racine):
        for fichier in fichiers:
            if fichier == 'Green_with_filled_contour.png':
                conn = sqlite3.connect('désertification.db')
                cur = conn.cursor()
                chemin_image = os.path.join(dossier, fichier)
                superficie_bleue_image = compter_pixels_bleus(chemin_image)
                superficie_bleue_image_km2 = (superficie_bleue_image * superficie_reference_km2)/pixel_bleu_reference
                volume = superficie_bleue_image_km2*1.6
                # Assurez-vous  que les données que vous ajoutez sont correctes
                print([os.path.basename(os.path.dirname(chemin_image)), superficie_bleue_image, superficie_bleue_image_km2])
                DATA = (str(chemin_image),str(sous_dossiers),int(superficie_bleue_image_km2),"1.6",int(volume))
                print(DATA)

                cur.execute("""INSERT INTO DESERTIFICATION (chemin, nom, surface, hauteur, volume)
                    VALUES(?, ?, ?, ?, ?)""", DATA)
                conn.commit()
                conn.close
                # Ajouter les données à la feuille de calcul
                ws.append([os.path.basename(os.path.dirname(chemin_image)), superficie_bleue_image, superficie_bleue_image_km2])
                # Enregistrement du classeur Excel

    # Enregistrement du classeur Excel
    wb.save(repertoire_script + "\Etude_de_la_flotte.xlsx")
    fichier_excel = os.path.join(repertoire_script, 'Etude_de_la_flotte.xlsx')
    main(fichier_excel)

    # Charger les deux tableaux
    arbres_df = pd.read_excel(os.path.join(repertoire_script, "marque_de_flotte.xlsx"))
    surfaces_df = pd.read_excel(os.path.join(repertoire_script, 'Etude_de_la_flotte.xlsx'))

    # Créer un DataFrame pour les résultats
    result_df = pd.DataFrame(columns=['Nom de l\'arbre', 'Noms associés', 'Image', 'Résultat (€/km²)', 'Résultat Hauteur (km*m²)', 'Nb arbre total'])

    # Parcourir les données des deux tableaux
    for idx_arbre, row_arbre in arbres_df.iterrows():
        for idx_surface, row_surface in surfaces_df.iterrows():
            # Calculer le résultat €/km²
            resultat = row_arbre['Prix au litre'] * row_surface['Surface en km²'] * 1600
            Nombretotal = row_arbre['Nb_kmcube_au_km2'] * row_surface['Surface en km²'] 

    # Enregistrez les résultats dans un fichier Excel
    result_df.to_excel(os.path.join(repertoire_script, 'Surfaces_au_complet.xlsx'), index=False)

    chemins_images_trouves = trouver_chemins_images(dossier_a_parcourir)

    enregistrer_chemins_dans_fichier(chemins_images_trouves, fichier_sortie)

    modifier_chemins_dans_fichier(fichier_sortie, fichier_sortie, repertoire_script)

    print(f"Les chemins des images ont été enregistrés dans {fichier_sortie}.")
    
    modifier_fichier(os.path.join(repertoire_script, 'interface', 'chemin.txt'), os.path.join(repertoire_script, 'interface', 'chemin.txt'))

    # Appelez la fonction pour ouvrir le fichier HTML
    ouvrir_html(chemin_fichier_html)
