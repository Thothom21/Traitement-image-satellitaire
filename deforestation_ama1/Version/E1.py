import tkinter as tk
import openpyxl

# Fonction pour lire les noms à partir du fichier Excel
def lire_noms_de_excel():
    noms = []
    wb = openpyxl.load_workbook('essence_darbres.xlsx')
    sheet = wb.active
    for cell in sheet['A'][1:]:  # Commence à partir de la deuxième ligne pour éviter l'en-tête
        if cell.value is not None:
            noms.append(cell.value)
    wb.close()
    return noms

# Fonction pour obtenir le numéro de ligne correspondant au nom sélectionné
def obtenir_numero_ligne(noms, nom_selectionne):
    try:
        return noms.index(nom_selectionne) + 2  # Ajouter 2 pour compenser l'index 0 et l'en-tête
    except ValueError:
        return None

# Fonction pour obtenir les données de la ligne correspondant au nom sélectionné
def obtenir_donnees_ligne(nom_selectionne):
    numero_ligne = obtenir_numero_ligne(noms, nom_selectionne)
    if numero_ligne is not None:
        ligne = list(sheet.iter_rows(min_row=numero_ligne, max_row=numero_ligne, values_only=True))[0]
        return ligne
    else:
        return None

# Fonction pour afficher les données de la ligne dans la console
def afficher_donnees_ligne():
    selected_nom = combo_box_var.get()
    entete_ligne = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    donnees_ligne = obtenir_donnees_ligne(selected_nom)
    if donnees_ligne is not None:
        print(f"Entête de la ligne : {entete_ligne}")
        print(f"Données de la ligne pour le nom '{selected_nom}': {donnees_ligne}")
    else:
        print("Nom non trouvé")

# Fonction pour quitter l'application
def quitter_application():
    root.destroy()  # Ferme la fenêtre principale de l'application

# Création de la fenêtre principale
root = tk.Tk()
root.title("Sélection de Nom")

# Lecture des noms depuis le fichier Excel
noms = lire_noms_de_excel()

# Création de la liste déroulante
combo_box_var = tk.StringVar()
combo_box_var.set(noms[0])  # Sélectionnez le premier nom par défaut

combo_box_label = tk.Label(root, text="Sélectionnez un nom :")
combo_box_label.pack()

combo_box = tk.OptionMenu(root, combo_box_var, *noms)
combo_box.pack()

# Ouverture du fichier Excel une fois pour définir 'sheet'
wb = openpyxl.load_workbook('essence_darbres.xlsx')
sheet = wb.active
wb.close()

# Bouton pour valider la sélection
valider_bouton = tk.Button(root, text="Valider", command=afficher_donnees_ligne)
valider_bouton.pack()

# Bouton "Quitter"
quitter_bouton = tk.Button(root, text="Quitter", command=quitter_application)
quitter_bouton.pack()

# Exécution de la boucle principale de l'interface utilisateur
root.mainloop()
