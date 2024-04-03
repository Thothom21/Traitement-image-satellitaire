import tkinter as tk
import openpyxl
from tkinter import messagebox

# Fonction pour lire les noms à partir du fichier Excel
def lire_noms_de_excel():
    noms = []
    wb = openpyxl.load_workbook('essence_darbres.xlsx')
    sheet = wb.active
    for cell in sheet['A'][1:]:  # Commence à partir de la deuxième ligne pour éviter l'en-tête
        if cell.value is not None:
            noms.append(cell.value)
    wb.close()
    return noms

# Fonction pour obtenir le numéro de ligne correspondant au nom sélectionné
def obtenir_numero_ligne(noms, nom_selectionne):
    try:
        return noms.index(nom_selectionne) + 2  # Ajouter 2 pour compenser l'index 0 et l'en-tête
    except ValueError:
        return None

# Fonction pour obtenir les données de la ligne correspondant au nom sélectionné
def obtenir_donnees_ligne(nom_selectionne):
    numero_ligne = obtenir_numero_ligne(noms, nom_selectionne)
    if numero_ligne is not None:
        ligne = list(sheet.iter_rows(min_row=numero_ligne, max_row=numero_ligne, values_only=True))[0]
        return ligne
    else:
        return None

# Fonction pour afficher les données de la ligne dans la console
def afficher_donnees_ligne():
    selected_nom = combo_box_var.get()
    donnees_ligne = obtenir_donnees_ligne(selected_nom)  # Obtenez les données de la ligne sélectionnée
    if donnees_ligne is not None:
        print(f"Données de la ligne pour le nom '{selected_nom}': {donnees_ligne}")
    else:
        print("Nom non trouvé")
import openpyxl
from tkinter import messagebox

# Fonction pour calculer le coût pour toutes les colonnes
def calculer_cout_pour_toutes_colonnes():
    selected_nom = combo_box_var.get()
    donnees_ligne = obtenir_donnees_ligne(selected_nom)  # Obtenez les données de la ligne sélectionnée

    if donnees_ligne is not None:
        # Ouvrir le fichier Excel "Etude_déforestation.xlsx"
        wb_etude = openpyxl.load_workbook('Etude_déforestation.xlsx')
        sheet_etude = wb_etude.active

        # Vérifier si la colonne "Coût" existe déjà, sinon la créer
        entete_liste = list(sheet_etude[1])
        if "Coût" not in entete_liste:
            entete_liste.insert(0, "Coût")  # Insérer "Coût" en première position
            for col, value in enumerate(entete_liste, 1):
                sheet_etude.cell(row=1, column=col, value=value)

        index_cout = entete_liste.index("Coût") + 1

        # Rechercher l'index de la colonne "Surface en km²" dans la première ligne
        index_surface = None
        for col, cell in enumerate(entete_liste, 1):
            if cell == "Surface en km²":
                index_surface = col
                break

        if index_surface is not None:
            # Parcourir les lignes à partir de la deuxième (données)
            for row in range(2, sheet_etude.max_row + 1):
                surface_km2 = sheet_etude.cell(row=row, column=index_surface).value
                prix_reforestation = donnees_ligne[entete_liste.index("Prix de la reforestation au km2")]
                cout = surface_km2 * prix_reforestation
                sheet_etude.cell(row=row, column=index_cout, value=cout)

        # Enregistrez le fichier Excel
        wb_etude.save('Etude_déforestation.xlsx')
        wb_etude.close()

        messagebox.showinfo("Calcul de Coût", "Le coût a été calculé pour toutes les colonnes et enregistré dans 'Etude_déforestation.xlsx'.")
    else:
        messagebox.showerror("Erreur", "Nom non trouvé ou données manquantes.")

# Fonction pour quitter l'application
def quitter_application():
    root.destroy()  # Ferme la fenêtre principale de l'application

# Création de la fenêtre principale
root = tk.Tk()
root.title("Sélection de Nom")

# Lecture des noms depuis le fichier Excel
noms = lire_noms_de_excel()

# Création de la liste déroulante
combo_box_var = tk.StringVar()
combo_box_var.set(noms[0])  # Sélectionnez le premier nom par défaut

combo_box_label = tk.Label(root, text="Sélectionnez un nom :")
combo_box_label.pack()

combo_box = tk.OptionMenu(root, combo_box_var, *noms)
combo_box.pack()

# Ouverture du fichier Excel une fois pour définir 'sheet'
wb = openpyxl.load_workbook('essence_darbres.xlsx')
sheet = wb.active
entete_ligne = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]  # Charger l'en-tête une fois
wb.close()

# Bouton pour valider la sélection
valider_bouton = tk.Button(root, text="Valider", command=afficher_donnees_ligne)
valider_bouton.pack()

# Bouton "Calculer Coût pour Toutes les Colonnes"
calculer_cout_toutes_colonnes_bouton = tk.Button(root, text="Calculer Coût pour Toutes les Colonnes", command=calculer_cout_pour_toutes_colonnes)
calculer_cout_toutes_colonnes_bouton.pack()

# Bouton "Quitter"
quitter_bouton = tk.Button(root, text="Quitter", command=quitter_application)
quitter_bouton.pack()

# Exécution de la boucle principale de l'interface utilisateur
root.mainloop()
