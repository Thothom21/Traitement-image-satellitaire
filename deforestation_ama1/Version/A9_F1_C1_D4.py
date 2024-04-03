import os
import cv2
import numpy as np
from openpyxl import Workbook
import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import Listbox, Text
import webbrowser

# Création d'un classeur Excel
wb = Workbook()
ws = wb.active

def modifier_fichier(input_path, output_path):
    with open(input_path, 'r') as fichier_entree:
        lignes = fichier_entree.readlines()

    lignes_modifiees = ['<img src="{}">'.format(ligne.strip()) for ligne in lignes]

    with open(output_path, 'w') as fichier_sortie:
        fichier_sortie.write('\n'.join(lignes_modifiees))


def trouver_chemins_images(dossier, extensions=('jpg', 'jpeg', 'png', 'gif')):
    chemins_images = []
    for dossier_actuel, sous_dossiers, fichiers in os.walk(dossier):
        for fichier in fichiers:
            if fichier.lower().endswith(extensions):
                chemin_complet = os.path.join(dossier_actuel, fichier)
                chemins_images.append(chemin_complet)
    return chemins_images

def enregistrer_chemins_dans_fichier(chemins_images, fichier_sortie='chemins_images.txt'):
    with open(fichier_sortie, 'w') as fichier:
        for chemin in chemins_images:
            fichier.write(chemin + '\n')

# Remplacez 'chemin_du_dossier' par le chemin du dossier que vous souhaitez parcourir
dossier_a_parcourir = r'C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\seuillage'

# Remplacez 'chemins_images.txt' par le chemin complet du fichier de sortie si nécessaire
fichier_sortie = r'C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\interface\chemin.txt'

# Ajout des entêtes
ws.append(["Nom de l'image", "Proportion de blanc (%)", "Surface en km²"])

def ouvrir_html(chemin_fichier):
    # Ouvrir le fichier HTML dans le navigateur par défaut
    webbrowser.open('file://' + chemin_fichier)

# Spécifiez le chemin complet de votre fichier HTML
chemin_fichier_html = r'C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\interface\testsupp.html'


def process_images_in_folder(input_folder, output_folder,lower_green,upper_green):
    # Create the output folders if they don't exist
    os.makedirs(output_folder, exist_ok=True)
    
    # Get a list of all image files in the input folder
    image_files = [f for f in os.listdir(input_folder) if f.endswith(('.jpg', '.png', '.jpeg', '.bmp', '.PNG'))]

    for image_file in image_files:
        # Load the image
        image_path = os.path.join(input_folder, image_file)
        image = cv2.imread(image_path)
        if image is None:
            print(f"Failed to read {image_file}. Skipping...")
            continue

        print(f"Processing {image_file}...")

        # Convertir l'image en format HSV (teinte, saturation, valeur)
        hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

        # Définir la plage de couleurs verte que vous souhaitez détecter
        lower_green = np.array(lower_green)
        upper_green = np.array(upper_green)

        # Créer un masque pour les pixels verts dans la plage spécifiée
        mask = cv2.inRange(hsv_image, lower_green, upper_green)

        # Appliquer le masque à l'image d'origine
        result = cv2.bitwise_and(image, image, mask=mask)

        # Convert to grayscale
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Save the grayscale image
        output_path = os.path.join(output_folder, image_file)
        os.makedirs(output_path, exist_ok=True)
        cv2.imwrite(os.path.join(output_path, "gray_image.png"), gray)

        # Apply Gaussian blur
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)

        # Apply Canny edge detection with different thresholds
        large = cv2.Canny(blurred, 10, 200)
        medium = cv2.Canny(blurred, 30, 150)
        narrow = cv2.Canny(blurred, 240, 250)

        # Save the processed images in the output folder
        cv2.imwrite(os.path.join(output_path, "large_canny.png"), large)
        cv2.imwrite(os.path.join(output_path, "medium_canny.png"), medium)
        cv2.imwrite(os.path.join(output_path, "narrow_canny.png"), narrow)
        cv2.imwrite(os.path.join(output_path, "Green_and_Black.png"), result)

        print(f"Processed {image_file} and saved the results in {output_path}")

# Fonction pour détecter les contours blancs et remplir les zones
def detecter_contours_et_remplir(image_path):
    image = cv2.imread(image_path)
    if image is None:
        return

    # Conversion en niveaux de gris
    image_gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Détection des contours blancs
    _, thresh = cv2.threshold(image_gray, 200, 255, cv2.THRESH_BINARY)

    # Dilatation
    kernel = np.ones((3, 3), np.uint8)
    thresh = cv2.dilate(thresh, kernel, iterations=5)

    # Érosion pour fermer les zones (moins agressivement)
    thresh = cv2.erode(thresh, kernel, iterations=5)

    # Trouver les contours
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # Créer une image vide pour le remplissage
    filled_image = np.zeros_like(image)

    # Remplir les contours en blanc
    cv2.drawContours(filled_image, contours, -1, (255, 255, 255), thickness=cv2.FILLED)

    # Enregistrement de l'image modifiée
    output_path = image_path.replace('medium_canny.png', 'filled.png')
    cv2.imwrite(output_path, filled_image)

    return filled_image

# Fonction pour compter les pixels et calculer la superficie en km²
def compter_pixels_et_calculer_superficie(image):
    if image is None:
        return None

    # Compter le nombre total de pixels
    total_pixels = image.size

    # Compter le nombre de pixels blancs
    white_pixels = np.count_nonzero(np.all(image == [255, 255, 255], axis=2))

    # Calculer la superficie en km²
    superficie_km2 = (white_pixels / (38 * 38)) * 400_000_000

    return total_pixels, white_pixels, superficie_km2

# Répertoire racine contenant les sous-dossiers d'images
repertoire_racine = r"C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\seuillage"

# Fonction pour supprimer l'extension d'un nom de fichier
def supprimer_extension(nom_fichier):
    if '.' in nom_fichier:
        return nom_fichier.split('.')[0]
    else:
        return nom_fichier

# Fonction principale
def main(nom_fichier):
    try:
        # Ouvrir le fichier Excel
        workbook = openpyxl.load_workbook(nom_fichier)
        sheet = workbook.active

        # Remplacer la cellule A1 par "Image"
        sheet['A1'] = "Image"

        # Parcourir les cellules de la colonne 1 et supprimer l'extension du contenu
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                contenu_cellule = cell.value
                if contenu_cellule:
                    cell.value = supprimer_extension(contenu_cellule)

        # Sauvegarder le fichier Excel modifié
        workbook.save(nom_fichier)
        print("Le fichier Excel a été modifié avec succès.")

    except Exception as e:
        print(f"Une erreur s'est produite : {str(e)}")


# Fonction pour afficher les résultats
def afficher_resultats():
    essence_selectionnee = liste_essences.get(liste_essences.curselection())
    resultat_text.delete(1.0, tk.END)

    for i in range(len(result_df)):
        if essence_selectionnee in result_df['Noms associés'][i].split(', '):
            resultat_text.insert(tk.END, f"Image: {result_df['Image'][i]}\n")
            resultat_text.insert(tk.END, f"Résultat (€/km²): {result_df['Résultat (€/km²)'][i]}\n")
            resultat_text.insert(tk.END, f"Résultat Hauteur (km*m²): {result_df['Résultat Hauteur (km*m²)'][i]}\n")
            resultat_text.insert(tk.END, f"Nb_arbre_total: {result_df['Nb_arbre_total'][i]}\n\n")




if __name__ == "__main__":
    input_folder = r"C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\photo"
    output_folder = r"C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\seuillage"
    
    # Plage de couleurs verte en format HSV (teinte, saturation, valeur)
    lower_green = [35, 50, 50]  # Plage inférieure (vert)
    upper_green = [85, 255, 255]  # Plage supérieure (vert)
    
    process_images_in_folder(input_folder, output_folder,lower_green,upper_green)

    # Parcourez les sous-dossiers et recherchez "medium_canny.png"
    for dossier, sous_dossiers, fichiers in os.walk(repertoire_racine):
        for fichier in fichiers:
            if fichier == 'medium_canny.png':
                chemin_image = os.path.join(dossier, fichier)
                filled_image = detecter_contours_et_remplir(chemin_image)
                if filled_image is not None:
                    total_pixels, white_pixels, superficie_km2 = compter_pixels_et_calculer_superficie(filled_image)
                    print(f"Image traitée : {chemin_image}")
                    print(f"Nombre total de pixels : {total_pixels}")
                    print(f"Nombre de pixels blancs : {white_pixels}")
                    print(f"Superficie en km² : {superficie_km2} km²")
                    # Utilisez le nom du dossier au lieu du nom de l'image
                    ws.append([os.path.basename(os.path.dirname(chemin_image)), (white_pixels / total_pixels) * 100, superficie_km2])

    # Enregistrement du classeur Excel
    wb.save("Etude_déforestation.xlsx")
    fichier_excel = r"C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\Etude_déforestation.xlsx"  # Remplacez par le nom de votre fichier Excel
    main(fichier_excel)


    # Charger les deux tableaux
    arbres_df = pd.read_excel(r"C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\Essence_d'arbres.xlsx")
    surfaces_df = pd.read_excel(r"C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\Etude_déforestation.xlsx")

    # Créer un DataFrame pour les résultats
    result_df = pd.DataFrame(columns=['Nom de l\'arbre', 'Noms associés', 'Image', 'Résultat (€/km²)', 'Résultat Hauteur (km*m²)', 'Nb arbre total'])

    # Parcourir les données des deux tableaux
    for idx_arbre, row_arbre in arbres_df.iterrows():
        for idx_surface, row_surface in surfaces_df.iterrows():
            # Calculer le résultat €/km²
            resultat = row_arbre['Prix km2'] * row_surface['Surface en km²']

            # Calculer le résultat de la hauteur (km*m²)
            resultat_hauteur = row_arbre['Hauteur_linéaire_arbre_km2*m'] * row_surface['Surface en km²']

            # Calculer le nombre total d'arbres
            Nombretotal = row_arbre['Nb_arbre_au_ km2'] * row_surface['Surface en km²']

            result_df = result_df._append({
                'Nom de l\'arbre': row_arbre['Nom'],
                'Noms associés': ', '.join(arbres_df[arbres_df['Nom'] == row_arbre['Nom']]['Nom'].tolist()),
                'Image': row_surface['Image'],
                'Résultat (€/km²)': resultat,
                'Résultat Hauteur (km*m²)': resultat_hauteur,
                'Nb_arbre_total': Nombretotal
            }, ignore_index=True)

    # Enregistrez les résultats dans un fichier Excel
    result_df.to_excel(r"C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\Surfaces_au_complet.xlsx", index=False)

    # Créer l'interface tkinter
    fenetre = tk.Tk()
    fenetre.title("Affichage des résultats par essence d'arbres")

    liste_essences = Listbox(fenetre)
    liste_essences.pack()
    essences_uniques = list(arbres_df['Nom'].unique())
    for essence in essences_uniques:
        liste_essences.insert(tk.END, essence)

    afficher_button = tk.Button(fenetre, text="Afficher les résultats", command=afficher_resultats)
    afficher_button.pack()

    resultat_text = Text(fenetre, height=40, width=70)
    resultat_text.pack()

    fenetre.mainloop()
    chemins_images_trouves = trouver_chemins_images(dossier_a_parcourir)
    enregistrer_chemins_dans_fichier(chemins_images_trouves, fichier_sortie)

    print(f"Les chemins des images ont été enregistrés dans {fichier_sortie}.")
    
    modifier_fichier(r'C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\interface\chemin.txt', r'C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\interface\chemin.txt')

    # Appelez la fonction pour ouvrir le fichier HTML
    ouvrir_html(chemin_fichier_html)
