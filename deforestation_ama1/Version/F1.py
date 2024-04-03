import openpyxl

# Fonction pour supprimer l'extension d'un nom de fichier
def supprimer_extension(nom_fichier):
    if '.' in nom_fichier:
        return nom_fichier.split('.')[0]
    else:
        return nom_fichier

# Fonction principale
def main(nom_fichier):
    try:
        # Ouvrir le fichier Excel
        workbook = openpyxl.load_workbook(nom_fichier)
        sheet = workbook.active

        # Remplacer la cellule A1 par "Image"
        sheet['A1'] = "Image"

        # Parcourir les cellules de la colonne 1 et supprimer l'extension du contenu
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                contenu_cellule = cell.value
                if contenu_cellule:
                    cell.value = supprimer_extension(contenu_cellule)

        # Sauvegarder le fichier Excel modifié
        workbook.save(nom_fichier)
        print("Le fichier Excel a été modifié avec succès.")

    except Exception as e:
        print(f"Une erreur s'est produite : {str(e)}")

if __name__ == "__main__":
    fichier_excel = r"C:\Users\bauma\OneDrive\Bureau\esirem_dijon\traitement\projet\Etude_déforestation.xlsx"  # Remplacez par le nom de votre fichier Excel
    main(fichier_excel)
