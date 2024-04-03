import os
import cv2
import shutil
import unittest
import openpyxl
import webbrowser
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import Listbox, Text
from openpyxl import Workbook, load_workbook

"""
PS : je fait une extraction vers txt pour les renvoyez dans ma page html et pour un programme annexes qui na rien avoir avec ce projet_ci 
ce qui justifie le fait de faire une extraction txt.

"""

"""
Algorithme ProgrammePrincipal:
    1. Créer un classeur Excel (Workbook) et une feuille active (Worksheet).

    2. Appeler la fonction `process_images_in_folder` avec les paramètres appropriés pour traiter les images.
         a. Créer les dossiers de sortie si nécessaire.
         b. Obtenir la liste de tous les fichiers d'image dans le dossier d'entrée.
         c. Pour chaque fichier d'image:
              i. Charger l'image.
              ii. Appliquer le seuillage pour détecter les pixels verts.
              iii. Enregistrer différentes versions de l'image traitée.

    3. Parcourir les sous-dossiers et rechercher les fichiers "medium_canny.png".
         a. Appeler la fonction `detecter_contours_et_remplir` pour détecter les contours et remplir les zones.
         b. Si réussi, appeler la fonction `compter_pixels_et_calculer_superficie` pour calculer les informations de superficie.
         c. Ajouter les résultats au classeur Excel avec des entêtes appropriées.

    4. Sauvegarder le classeur Excel avec le nom "Etude_déforestation.xlsx".

    5. Charger les données des fichiers Excel (arbres et surfaces) dans des DataFrames.

    6. Créer un DataFrame pour stocker les résultats finaux.

    7. Pour chaque ligne d'arbres et chaque ligne de surfaces:
         a. Calculer le résultat €/km², le résultat de la hauteur (km*m²) et le nombre total d'arbres.
         b. Ajouter ces informations au DataFrame des résultats.

    8. Enregistrer les résultats dans un nouveau fichier Excel nommé "Surfaces_au_complet.xlsx".

    9. Créer une interface tkinter pour afficher les résultats par essence d'arbres.

    10. Afficher les résultats lorsque l'utilisateur sélectionne une essence et clique sur le bouton.

    11. Enregistrer les chemins des images dans un fichier "chemins_images.txt".

    12. Modifier les chemins pour les passer en relatif 

    13. Modifier un fichier HTML en utilisant la fonction `modifier_fichier`.

    14. Ouvrir le fichier HTML dans le navigateur par défaut.

Fin de l'algorithme.

"""

####################################################################################################################
def modifier_fichier(input_path, output_path):
    """
    Lit un fichier d'entrée, ajoute des balises HTML <img> autour de chaque ligne,
    puis écrit le résultat dans un fichier de sortie.

    Args:
        input_path (str): Le chemin du fichier d'entrée.
        output_path (str): Le chemin du fichier de sortie.
    """
    with open(input_path, 'r') as fichier_entree:
        lignes = fichier_entree.readlines()

    lignes_modifiees = ['<img src="..\{}">'.format(ligne.strip()) for ligne in lignes]

    with open(output_path, 'w') as fichier_sortie:
        fichier_sortie.write('\n'.join(lignes_modifiees))

# Doctest pour la fonction modifier_fichier
def test_modifier_fichier():
    """
    >>> input_path_test = 'test_input.txt'
    >>> output_path_test = 'test_output.txt'
    >>> with open(input_path_test, 'w') as f:
    ...     f.write('image1.jpg\\nimage2.png\\nimage3.gif')
    >>> modifier_fichier(input_path_test, output_path_test)
    >>> with open(output_path_test, 'r') as f:
    ...     modified_content = f.read()
    >>> print(modified_content)
    <img src="image1.jpg">
    <img src="image2.png">
    <img src="image3.gif">
    >>> os.remove(input_path_test)
    >>> os.remove(output_path_test)
    """
    pass

####################################################################################################################

def trouver_chemins_images(dossier, extensions=('jpg', 'jpeg', 'png', 'gif')):
    """
    Recherche récursivement tous les fichiers d'images dans un dossier.

    Args:
        dossier (str): Le chemin du dossier à parcourir.
        extensions (tuple): Les extensions de fichiers à rechercher. Par défaut, ('jpg', 'jpeg', 'png', 'gif').

    Returns:
        List[str]: Une liste contenant les chemins complets de tous les fichiers d'images trouvés.
    """
    chemins_images = []
    for dossier_actuel, sous_dossiers, fichiers in os.walk(dossier):
        for fichier in fichiers:
            if fichier.lower().endswith(extensions):
                chemin_complet = os.path.join(dossier_actuel, fichier)
                chemins_images.append(chemin_complet)
    return chemins_images

# Doctest pour la fonction trouver_chemins_images
def test_trouver_chemins_images():
    """
    >>> import tempfile
    >>> tmpdir = tempfile.mkdtemp()
    >>> img1_path = os.path.join(tmpdir, 'image1.jpg')
    >>> img2_path = os.path.join(tmpdir, 'image2.png')
    >>> img3_path = os.path.join(tmpdir, 'subfolder', 'image3.gif')
    >>> os.makedirs(os.path.join(tmpdir, 'subfolder'))
    >>> open(img1_path, 'w').close()
    >>> open(img2_path, 'w').close()
    >>> open(img3_path, 'w').close()

    >>> chemins = trouver_chemins_images(tmpdir)
    >>> len(chemins)
    3
    >>> img1_path in chemins
    True
    >>> img2_path in chemins
    True
    >>> img3_path in chemins
    True

    >>> img4_path = os.path.join(tmpdir, 'document.txt')
    >>> open(img4_path, 'w').close()
    >>> chemins = trouver_chemins_images(tmpdir)
    >>> len(chemins)
    3
    >>> img4_path in chemins
    False

    >>> img5_path = os.path.join(tmpdir, 'image5.jpg')
    >>> os.makedirs(os.path.join(tmpdir, 'empty_folder'))
    >>> open(img5_path, 'w').close()
    >>> chemins = trouver_chemins_images(os.path.join(tmpdir, 'empty_folder'))
    >>> len(chemins)
    0

    >>> import shutil
    >>> shutil.rmtree(tmpdir)
    """
    pass

####################################################################################################################

def enregistrer_chemins_dans_fichier(chemins_images, fichier_sortie='chemins_images.txt'):
    """
    Enregistre les chemins des images dans un fichier texte.

    Args:
        chemins_images (List[str]): Une liste contenant les chemins complets des fichiers d'images.
        fichier_sortie (str): Le chemin du fichier de sortie. Par défaut, 'chemins_images.txt'.
    """
    with open(fichier_sortie, 'w') as fichier:
        for chemin in chemins_images:
            fichier.write(chemin + '\n')

# Doctest pour la fonction enregistrer_chemins_dans_fichier
def test_enregistrer_chemins_dans_fichier():
    """
    >>> import tempfile
    >>> tmpdir = tempfile.mkdtemp()
    >>> img1_path = os.path.join(tmpdir, 'image1.jpg')
    >>> img2_path = os.path.join(tmpdir, 'image2.png')
    >>> img3_path = os.path.join(tmpdir, 'subfolder', 'image3.gif')
    >>> os.makedirs(os.path.join(tmpdir, 'subfolder'))
    >>> open(img1_path, 'w').close()
    >>> open(img2_path, 'w').close()
    >>> open(img3_path, 'w').close()

    >>> chemins = [img1_path, img2_path, img3_path]
    >>> fichier_sortie = os.path.join(tmpdir, 'output.txt')
    >>> enregistrer_chemins_dans_fichier(chemins, fichier_sortie)

    >>> with open(fichier_sortie, 'r') as f:
    ...     contenu_fichier = f.read()
    >>> print(contenu_fichier)
    {0}
    {1}
    {2}

    >>> shutil.rmtree(tmpdir)
    """
    pass

####################################################################################################################


def modifier_chemins_dans_fichier(input_path, output_path, dossier_base):
    """
    Modifie les chemins absolus dans un fichier pour les rendre relatifs par rapport à un dossier de base.

    Args:
        input_path (str): Chemin du fichier d'entrée contenant les chemins absolus.
        output_path (str): Chemin du fichier de sortie pour enregistrer les chemins relatifs.
        dossier_base (str): Dossier de base pour rendre les chemins relatifs.
    """
    with open(input_path, 'r') as fichier_entree:
        lignes = fichier_entree.readlines()

    chemins_relatifs = [os.path.relpath(chemin.strip(), dossier_base) for chemin in lignes]

    with open(output_path, 'w') as fichier_sortie:
        fichier_sortie.write('\n'.join(chemins_relatifs))


# Doctest pour la fonction modifier_chemins_dans_fichier
def test_modifier_chemins_dans_fichier():
    """
    >>> import tempfile
    >>> tmpdir = tempfile.mkdtemp()
    >>> img1_path = os.path.join(tmpdir, 'image1.jpg')
    >>> img2_path = os.path.join(tmpdir, 'image2.png')
    >>> img3_path = os.path.join(tmpdir, 'subfolder', 'image3.gif')
    >>> os.makedirs(os.path.join(tmpdir, 'subfolder'))
    >>> open(img1_path, 'w').close()
    >>> open(img2_path, 'w').close()
    >>> open(img3_path, 'w').close()

    >>> chemins_absolus = [img1_path, img2_path, img3_path]
    >>> fichier_entrée = os.path.join(tmpdir, 'input.txt')
    >>> fichier_sortie = os.path.join(tmpdir, 'output.txt')
    
    >>> with open(fichier_entrée, 'w') as f:
    ...     f.write('\\n'.join(chemins_absolus))

    >>> dossier_base = tmpdir
    >>> modifier_chemins_dans_fichier(fichier_entrée, fichier_sortie, dossier_base)

    >>> with open(fichier_sortie, 'r') as f:
    ...     chemins_relatifs = f.read().splitlines()
    
    >>> print(chemins_relatifs)
    ['image1.jpg', 'image2.png', 'subfolder/image3.gif']

    >>> shutil.rmtree(tmpdir)
    """
    pass



####################################################################################################################

def ouvrir_html(chemin_fichier):
    """
    Ouvre le fichier HTML dans le navigateur par défaut.

    Args:
        chemin_fichier (str): Le chemin complet du fichier HTML.

    Example:
        >>> import tempfile
        >>> import webbrowser
        >>> tmpdir = tempfile.mkdtemp()
        >>> fichier_html = os.path.join(tmpdir, 'test.html')
        >>> open(fichier_html, 'w').close()
        >>> ouvrir_html(fichier_html)
        >>> # Note: La fonction webbrowser.open ne renvoie pas de valeur directe,
        >>> # mais elle devrait ouvrir le fichier dans le navigateur par défaut.

        >>> shutil.rmtree(tmpdir)
    """
    webbrowser.open('file://' + chemin_fichier)


####################################################################################################################

def process_images_in_folder(input_folder, output_folder, lower_green, upper_green):
    """
    Traite les images dans un dossier, applique la détection de couleur verte,
    crée des images en niveaux de gris, applique un flou gaussien,
    effectue une détection de contour Canny à différents seuils, et enregistre les résultats.

    Args:
        input_folder (str): Le chemin du dossier d'entrée contenant les images.
        output_folder (str): Le chemin du dossier de sortie pour enregistrer les résultats.
        lower_green (tuple): Les composantes HSV inférieures de la plage de vert à détecter.
        upper_green (tuple): Les composantes HSV supérieures de la plage de vert à détecter.
    """
    # Créer les dossiers de sortie s'ils n'existent pas
    os.makedirs(output_folder, exist_ok=True)
    
    # Obtenir la liste de tous les fichiers image dans le dossier d'entrée
    image_files = [f for f in os.listdir(input_folder) if f.endswith(('.jpg', '.png', '.jpeg', '.bmp', '.PNG'))]

    for image_file in image_files:
        # Charger l'image
        image_path = os.path.join(input_folder, image_file)
        image = cv2.imread(image_path)
        if image is None:
            print(f"Échec de la lecture de {image_file}. Ignorer...")
            continue

        print(f"Traitement de {image_file}...")

        # Convertir l'image en format HSV (teinte, saturation, valeur)
        hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

        # Définir la plage de couleurs verte que vous souhaitez détecter
        lower_green = np.array(lower_green)
        upper_green = np.array(upper_green)

        # Créer un masque pour les pixels verts dans la plage spécifiée
        mask = cv2.inRange(hsv_image, lower_green, upper_green)

        # Appliquer le masque à l'image d'origine
        result = cv2.bitwise_and(image, image, mask=mask)

        # Convertir en niveaux de gris
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Enregistrer l'image en niveaux de gris
        output_path = os.path.join(output_folder, image_file)
        os.makedirs(output_path, exist_ok=True)
        cv2.imwrite(os.path.join(output_path, "gray_image.png"), gray)

        # Appliquer un flou gaussien
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)

        # Appliquer la détection de contours Canny avec différents seuils
        large = cv2.Canny(blurred, 10, 200)
        medium = cv2.Canny(blurred, 30, 150)
        narrow = cv2.Canny(blurred, 240, 250)

        # Enregistrer les images traitées dans le dossier de sortie
        cv2.imwrite(os.path.join(output_path, "large_canny.png"), large)
        cv2.imwrite(os.path.join(output_path, "medium_canny.png"), medium)
        cv2.imwrite(os.path.join(output_path, "narrow_canny.png"), narrow)
        cv2.imwrite(os.path.join(output_path, "Green_and_Black.png"), result)

        print(f"{image_file} traité et les résultats ont été enregistrés dans {output_path}")

# Doctest pour la fonction process_images_in_folder
def test_process_images_in_folder():
    """
    >>> input_folder_test = 'test_input_images'
    >>> output_folder_test = 'test_output_images'
    >>> lower_green_test = (40, 40, 40)
    >>> upper_green_test = (80, 255, 255)
    >>> os.makedirs(input_folder_test, exist_ok=True)
    >>> os.makedirs(output_folder_test, exist_ok=True)
    >>> image_path_test = os.path.join(input_folder_test, 'test_image.jpg')
    >>> cv2.imwrite(image_path_test, np.zeros((100, 100, 3), dtype=np.uint8))  # Crée une image noire
    >>> process_images_in_folder(input_folder_test, output_folder_test, lower_green_test, upper_green_test)
    >>> # Vérifier si les fichiers résultants ont été créés dans le dossier de sortie
    >>> assert os.path.isfile(os.path.join(output_folder_test, 'test_image', 'gray_image.png'))
    >>> assert os.path.isfile(os.path.join(output_folder_test, 'test_image', 'large_canny.png'))
    >>> assert os.path.isfile(os.path.join(output_folder_test, 'test_image', 'medium_canny.png'))
    >>> assert os.path.isfile(os.path.join(output_folder_test, 'test_image', 'narrow_canny.png'))
    >>> assert os.path.isfile(os.path.join(output_folder_test, 'test_image', 'Green_and_Black.png'))
    >>> # Nettoyer les fichiers de test
    >>> os.remove(image_path_test)
    >>> os.rmdir(os.path.join(output_folder_test, 'test_image'))
    >>> os.rmdir(output_folder_test)
    >>> os.rmdir(input_folder_test)
    """
    pass

####################################################################################################################

def detecter_contours_et_remplir(image_path):
    """
    Détecte les contours blancs dans une image, remplit les zones encloses,
    et enregistre l'image résultante.

    Args:
        image_path (str): Le chemin de l'image à traiter.

    Returns:
        np.ndarray or None: L'image remplie en blanc ou None si la lecture de l'image échoue.
    """
    # Charger l'image
    image = cv2.imread(image_path)
    if image is None:
        return None

    # Conversion en niveaux de gris
    image_gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Détection des contours blancs
    _, thresh = cv2.threshold(image_gray, 200, 255, cv2.THRESH_BINARY)

    # Dilatation
    kernel = np.ones((3, 3), np.uint8)
    thresh = cv2.dilate(thresh, kernel, iterations=5)

    # Érosion pour fermer les zones (moins agressivement)
    thresh = cv2.erode(thresh, kernel, iterations=5)

    # Trouver les contours
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # Créer une image vide pour le remplissage
    filled_image = np.zeros_like(image)

    # Remplir les contours en blanc
    cv2.drawContours(filled_image, contours, -1, (255, 255, 255), thickness=cv2.FILLED)

    # Enregistrement de l'image modifiée
    output_path = image_path.replace('medium_canny.png', 'filled.png')
    cv2.imwrite(output_path, filled_image)

    return filled_image

# Doctest pour la fonction detecter_contours_et_remplir
def test_detecter_contours_et_remplir():
    """
    >>> input_image_path = 'test_image_medium_canny.png'
    >>> output_image_path = 'test_image_filled.png'
    >>> # Créer une image avec un contour blanc
    >>> image = np.zeros((100, 100, 3), dtype=np.uint8)
    >>> cv2.rectangle(image, (10, 10), (90, 90), (255, 255, 255), thickness=2)
    >>> cv2.imwrite(input_image_path, image)
    >>> filled_image = detecter_contours_et_remplir(input_image_path)
    >>> assert filled_image is not None
    >>> assert np.all(filled_image[15, 15] == [255, 255, 255])  # Vérifier le remplissage à l'intérieur du contour
    >>> os.remove(input_image_path)
    >>> os.remove(output_image_path)
    """
    pass

####################################################################################################################

def compter_pixels_et_calculer_superficie(image):
    """
    Compte le nombre total de pixels, le nombre de pixels blancs et calcule la superficie en km²
    pour une image donnée.

    Args:
        image (numpy.ndarray): L'image à analyser, représentée comme un tableau NumPy.

    Returns:
        tuple or None: Un tuple contenant le nombre total de pixels, le nombre de pixels blancs
        et la superficie en km², ou None si l'image est None.
    """
    if image is None:
        return None

    # Compter le nombre total de pixels
    total_pixels = image.size

    # Compter le nombre de pixels blancs
    white_pixels = np.count_nonzero(np.all(image == [255, 255, 255], axis=2))

    # Calculer la superficie en km²
    superficie_km2 = (white_pixels / (38 * 38)) * 400_000_000

    return total_pixels, white_pixels, superficie_km2

# Doctest pour la fonction compter_pixels_et_calculer_superficie
def test_compter_pixels_et_calculer_superficie():
    """
    >>> import numpy as np
    >>> # Créer une image de 100x100 pixels avec un carré blanc au centre
    >>> image = np.zeros((100, 100, 3), dtype=np.uint8)
    >>> cv2.rectangle(image, (30, 30), (70, 70), (255, 255, 255), thickness=cv2.FILLED)
    >>> total_pixels, white_pixels, superficie_km2 = compter_pixels_et_calculer_superficie(image)
    >>> assert total_pixels == 30000  # 100x100x3
    >>> assert white_pixels == 1600   # (70-30)x(70-30)
    >>> assert superficie_km2 == 4210526.315789474
    """
    pass
####################################################################################################################
def supprimer_extension(nom_fichier):
    """
    Supprime l'extension d'un nom de fichier.

    Args:
        nom_fichier (str): Le nom du fichier avec ou sans extension.

    Returns:
        str: Le nom du fichier sans extension, ou le même nom si aucune extension n'est présente.
    """
    if '.' in nom_fichier:
        return nom_fichier.split('.')[0]
    else:
        return nom_fichier

# Doctest pour la fonction supprimer_extension
def test_supprimer_extension():
    """
    >>> supprimer_extension('fichier.txt')
    'fichier'
    >>> supprimer_extension('image.png')
    'image'
    >>> supprimer_extension('document')
    'document'
    >>> supprimer_extension('archive.tar.gz')
    'archive'
    >>> supprimer_extension('sans_extension')
    'sans_extension'
    """
    pass
####################################################################################################################

def main(nom_fichier):
    """
    Modifie un fichier Excel en remplaçant le contenu de la colonne 1 par les noms de fichiers sans extension.

    Args:
        nom_fichier (str): Le chemin du fichier Excel à modifier.

    Raises:
        Exception: Une exception est levée en cas d'erreur lors du traitement du fichier Excel.

    Example:
        Pour utiliser cette fonction, appelez-la avec le chemin du fichier Excel :
        >>> main('chemin/vers/votre/fichier.xlsx')
        Le fichier Excel a été modifié avec succès.
    """
    try:
        # Ouvrir le fichier Excel
        workbook = openpyxl.load_workbook(nom_fichier)
        sheet = workbook.active

        # Remplacer la cellule A1 par "Image"
        sheet['A1'] = "Image"

        # Parcourir les cellules de la colonne 1 et supprimer l'extension du contenu
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                contenu_cellule = cell.value
                if contenu_cellule:
                    cell.value = supprimer_extension(contenu_cellule)

        # Sauvegarder le fichier Excel modifié
        workbook.save(nom_fichier)
        print("Le fichier Excel a été modifié avec succès.")

    except Exception as e:
        print(f"Une erreur s'est produite : {str(e)}")

# Doctest pour la fonction main
def test_main():
    """
    >>> import openpyxl
    >>> import os
    >>> file_path = 'test_file.xlsx'
    >>> workbook = openpyxl.Workbook()
    >>> sheet = workbook.active
    >>> sheet['A1'] = 'test_file.jpg'
    >>> workbook.save(file_path)
    >>> main(file_path)
    Le fichier Excel a été modifié avec succès.
    >>> loaded_workbook = openpyxl.load_workbook(file_path)
    >>> loaded_sheet = loaded_workbook.active
    >>> cell_value = loaded_sheet['A1'].value
    >>> assert cell_value == 'test_file'
    >>> os.remove(file_path)
    """
    pass

####################################################################################################################

def afficher_resultats():
    """
    Affiche les résultats dans une zone de texte en fonction de l'essence sélectionnée.

    Utilise les variables globales liste_essences, resultat_text et result_df.

    Example:
        Pour utiliser cette fonction, assurez-vous que les variables globales nécessaires sont définies :
        - liste_essences : Liste déroulante (tkinter) contenant les essences.
        - resultat_text : Zone de texte (tkinter) pour afficher les résultats.
        - result_df : DataFrame pandas contenant les résultats.

        >>> # Exemple d'utilisation
        >>> import tkinter as tk
        >>> import pandas as pd
        >>> root = tk.Tk()
        >>> liste_essences = tk.Listbox(root)
        >>> resultat_text = tk.Text(root)
        >>> result_df = pd.DataFrame({
        ...     'Image': ['image1.jpg', 'image2.png'],
        ...     'Noms associés': ['essence1', 'essence2'],
        ...     'Résultat (€/km²)': [100, 200],
        ...     'Résultat Hauteur (km*m²)': [50, 75],
        ...     'Nb_arbre_total': [30, 45]
        ... })
        >>> afficher_resultats()
        >>> # Vérifiez manuellement l'affichage dans la zone de texte.
        >>> root.destroy()
    """
    essence_selectionnee = liste_essences.get(liste_essences.curselection())
    resultat_text.delete(1.0, tk.END)

    for i in range(len(result_df)):
        if essence_selectionnee in result_df['Noms associés'][i].split(', '):
            resultat_text.insert(tk.END, f"Image: {result_df['Image'][i]}\n")
            resultat_text.insert(tk.END, f"Résultat (€/km²): {result_df['Résultat (€/km²)'][i]}\n")
            resultat_text.insert(tk.END, f"Résultat Hauteur (km*m²): {result_df['Résultat Hauteur (km*m²)'][i]}\n")
            resultat_text.insert(tk.END, f"Nb_arbre_total: {result_df['Nb_arbre_total'][i]}\n\n")

# Doctest pour la fonction afficher_resultats
def test_afficher_resultats():
    """
    >>> import tkinter as tk
    >>> import pandas as pd
    >>> root = tk.Tk()
    >>> liste_essences = tk.Listbox(root)
    >>> resultat_text = tk.Text(root)
    >>> result_df = pd.DataFrame({
    ...     'Image': ['image1.jpg', 'image2.png'],
    ...     'Noms associés': ['essence1', 'essence2'],
    ...     'Résultat (€/km²)': [100, 200],
    ...     'Résultat Hauteur (km*m²)': [50, 75],
    ...     'Nb_arbre_total': [30, 45]
    ... })
    >>> afficher_resultats()
    >>> # Vérifiez manuellement l'affichage dans la zone de texte.
    >>> root.destroy()
    """
    pass

####################################################################################################################

# Obtenez le répertoire actuel du script
repertoire_script = os.path.dirname(os.path.abspath(__file__))

# Remplacez 'chemin_du_dossier' par le chemin du dossier que vous souhaitez parcourir
dossier_a_parcourir = os.path.join(repertoire_script, 'seuillage')

# Remplacez 'chemins_images.txt' par le chemin complet du fichier de sortie si nécessaire
fichier_sortie = os.path.join(repertoire_script, 'interface', 'chemin.txt')

# Spécifiez le chemin complet de votre fichier HTML
chemin_fichier_html = os.path.join(repertoire_script, 'interface', 'testsupp.html')

# Répertoire racine contenant les sous-dossiers d'images
repertoire_racine = os.path.join(repertoire_script, 'seuillage')

# Création d'un classeur Excel
wb = Workbook()
ws = wb.active
# Ajout des entêtes
ws.append(["Nom de l'image", "Proportion de blanc (%)", "Surface en km²"])


if __name__ == "__main__":
    input_folder = os.path.join(repertoire_script, 'photo')
    output_folder = os.path.join(repertoire_script, 'seuillage')
    
    # Plage de couleurs verte en format HSV (teinte, saturation, valeur)
    lower_green = [35, 50, 50]  # Plage inférieure (vert)
    upper_green = [85, 255, 255]  # Plage supérieure (vert)
    
    process_images_in_folder(input_folder, output_folder,lower_green,upper_green)

    # Parcourez les sous-dossiers et recherchez "medium_canny.png"
    for dossier, sous_dossiers, fichiers in os.walk(repertoire_racine):
        for fichier in fichiers:
            if fichier == 'medium_canny.png':
                chemin_image = os.path.join(dossier, fichier)
                filled_image = detecter_contours_et_remplir(chemin_image)
                if filled_image is not None:
                    total_pixels, white_pixels, superficie_km2 = compter_pixels_et_calculer_superficie(filled_image)
                    print(f"Image traitée : {chemin_image}")
                    print(f"Nombre total de pixels : {total_pixels}")
                    print(f"Nombre de pixels blancs : {white_pixels}")
                    print(f"Superficie en km² : {superficie_km2} km²")
                    # Utilisez le nom du dossier au lieu du nom de l'image
                    ws.append([os.path.basename(os.path.dirname(chemin_image)), (white_pixels / total_pixels) * 100, superficie_km2])

    # Enregistrement du classeur Excel
    wb.save("Etude_déforestation.xlsx")
    fichier_excel = os.path.join(repertoire_script, 'Etude_déforestation.xlsx')
    main(fichier_excel)


    # Charger les deux tableaux
    arbres_df = pd.read_excel(os.path.join(repertoire_script, "Essence_d'arbres.xlsx"))
    surfaces_df = pd.read_excel(os.path.join(repertoire_script, 'Etude_déforestation.xlsx'))

    # Créer un DataFrame pour les résultats
    result_df = pd.DataFrame(columns=['Nom de l\'arbre', 'Noms associés', 'Image', 'Résultat (€/km²)', 'Résultat Hauteur (km*m²)', 'Nb arbre total'])

    # Parcourir les données des deux tableaux
    for idx_arbre, row_arbre in arbres_df.iterrows():
        for idx_surface, row_surface in surfaces_df.iterrows():
            # Calculer le résultat €/km²
            resultat = row_arbre['Prix km2'] * row_surface['Surface en km²']

            # Calculer le résultat de la hauteur (km*m²)
            resultat_hauteur = row_arbre['Hauteur_linéaire_arbre_km2*m'] * row_surface['Surface en km²']

            # Calculer le nombre total d'arbres
            Nombretotal = row_arbre['Nb_arbre_au_ km2'] * row_surface['Surface en km²']

            result_df = result_df._append({
                'Nom de l\'arbre': row_arbre['Nom'],
                'Noms associés': ', '.join(arbres_df[arbres_df['Nom'] == row_arbre['Nom']]['Nom'].tolist()),
                'Image': row_surface['Image'],
                'Résultat (€/km²)': resultat,
                'Résultat Hauteur (km*m²)': resultat_hauteur,
                'Nb_arbre_total': Nombretotal
            }, ignore_index=True)

    # Enregistrez les résultats dans un fichier Excel
    result_df.to_excel(os.path.join(repertoire_script, 'Surfaces_au_complet.xlsx'), index=False)

    # Créer l'interface tkinter
    fenetre = tk.Tk()
    fenetre.title("Affichage des résultats par essence d'arbres")

    liste_essences = Listbox(fenetre)
    liste_essences.pack()
    essences_uniques = list(arbres_df['Nom'].unique())
    for essence in essences_uniques:
        liste_essences.insert(tk.END, essence)

    afficher_button = tk.Button(fenetre, text="Afficher les résultats", command=afficher_resultats)
    afficher_button.pack()

    resultat_text = Text(fenetre, height=40, width=70)
    resultat_text.pack()

    fenetre.mainloop()
    chemins_images_trouves = trouver_chemins_images(dossier_a_parcourir)

    enregistrer_chemins_dans_fichier(chemins_images_trouves, fichier_sortie)

    modifier_chemins_dans_fichier(fichier_sortie, fichier_sortie, repertoire_script)

    print(f"Les chemins des images ont été enregistrés dans {fichier_sortie}.")
    
    modifier_fichier(os.path.join(repertoire_script, 'interface', 'chemin.txt'), os.path.join(repertoire_script, 'interface', 'chemin.txt'))

    # Appelez la fonction pour ouvrir le fichier HTML
    ouvrir_html(chemin_fichier_html)




#############TEST UNITAIRE###################################################

import unittest
import tempfile
import os
import unittest
import os
import cv2
import numpy as np
from unittest.mock import patch
import unittest
import os
import cv2
import numpy as np
from unittest.mock import MagicMock, patch

class TestImageUtils(unittest.TestCase):

    def test_modifier_fichier(self):
        # Créer un fichier d'entrée temporaire
        with tempfile.NamedTemporaryFile(mode='w', delete=False) as temp_input:
            temp_input.write("Contenu du fichier\n")
            temp_input.flush()

            # Créer un fichier de sortie temporaire
            with tempfile.NamedTemporaryFile(mode='w', delete=False) as temp_output:
                temp_output_name = temp_output.name

                # Appeler la fonction à tester
                modifier_fichier(temp_input.name, temp_output_name)

                # Vérifier si le fichier de sortie a été modifié correctement
                with open(temp_output_name, 'r') as modified_file:
                    modified_content = modified_file.read()
                    expected_content = '<img src="Contenu du fichier">\n'
                    self.assertEqual(modified_content, expected_content)

                # Supprimer les fichiers temporaires après les tests
                os.remove(temp_input.name)
                os.remove(temp_output_name)

    def test_trouver_chemins_images(self):
        
        # Créer un dossier temporaire avec des fichiers images simulés

        with tempfile.TemporaryDirectory() as tempdir:
            # Créer quelques fichiers simulés dans le dossier temporaire
            images_simulees = ['image1.jpg', 'image2.png', 'image3.gif']
            for image_simulee in images_simulees:
                open(os.path.join(tempdir, image_simulee), 'a').close()

            # Appeler la fonction à tester
            chemins_images = trouver_chemins_images(tempdir)

            # Vérifier si la fonction renvoie la liste attendue
            self.assertEqual(chemins_images, [os.path.join(tempdir, image) for image in images_simulees])

    def test_enregistrer_chemins_dans_fichier(self):
        # Créer une liste de chemins simulés
        chemins_simules = ['/chemin/image1.jpg', '/chemin/image2.png', '/chemin/image3.gif']

        # Créer un dossier temporaire et un fichier de sortie simulé
        import tempfile
        with tempfile.TemporaryDirectory() as tempdir:
            fichier_sortie = os.path.join(tempdir, 'chemins_simules.txt')

            # Appeler la fonction à tester
            enregistrer_chemins_dans_fichier(chemins_simules, fichier_sortie)

            # Vérifier si le fichier de sortie a été créé
            self.assertTrue(os.path.isfile(fichier_sortie))

            # Lire le contenu du fichier et vérifier s'il correspond aux chemins simulés
            with open(fichier_sortie, 'r') as fichier:
                chemins_enregistres = fichier.readlines()

            self.assertEqual(chemins_enregistres, [chemin + '\n' for chemin in chemins_simules])

    def test_ouvrir_html(self):
            # Créer un fichier HTML simulé
            with open('test_html_file.html', 'w') as html_file:
                html_file.write('<html><body><h1>Test HTML</h1></body></html>')

            # Appeler la fonction à tester avec le fichier HTML simulé
            with patch('webbrowser.open') as mock_open:
                ouvrir_html('test_html_file.html')

                # Vérifier si la fonction webbrowser.open a été appelée avec le bon argument
                mock_open.assert_called_once_with('file://' + os.path.abspath('test_html_file.html'))

            # Supprimer le fichier HTML simulé après le test
            os.remove('test_html_file.html')

    def test_process_images_in_folder(self):
        # Créer un dossier temporaire avec des fichiers images simulés
        import tempfile

        with tempfile.TemporaryDirectory() as tempdir:
            # Créer quelques fichiers simulés dans le dossier temporaire
            images_simulees = ['image1.jpg', 'image2.png', 'image3.bmp']
            for image_simulee in images_simulees:
                image_path = os.path.join(tempdir, image_simulee)
                cv2.imwrite(image_path, np.zeros((100, 100, 3), dtype=np.uint8))

            # Appeler la fonction à tester avec le dossier temporaire simulé
            with patch('builtins.print') as mock_print:
                process_images_in_folder(tempdir, 'output_folder', [30, 50, 50], [90, 255, 255])

                # Vérifier si la fonction a été appelée avec les bons arguments
                self.assertEqual(mock_print.call_count, len(images_simulees) * 2)

        # Supprimer le dossier temporaire après le test

    def test_detecter_contours_et_remplir(self):
        # Créer un fichier d'image simulé
        image_simulee = np.zeros((100, 100, 3), dtype=np.uint8)
        image_path = 'test_image.png'
        cv2.imwrite(image_path, image_simulee)

        # Appeler la fonction à tester avec l'image simulée
        filled_image = detecter_contours_et_remplir(image_path)

        # Vérifier si l'image remplie a été créée
        self.assertTrue(os.path.exists(image_path.replace('medium_canny.png', 'filled.png')))
        self.assertTrue(np.array_equal(filled_image, cv2.imread(image_path.replace('medium_canny.png', 'filled.png'))))

        # Supprimer le fichier image après le test
        os.remove(image_path)

    def test_compter_pixels_et_calculer_superficie(self):
        # Créer une image simulée avec des pixels blancs
        image_simulee = np.ones((100, 100, 3), dtype=np.uint8) * 255

        # Appeler la fonction à tester avec l'image simulée
        total_pixels, white_pixels, superficie_km2 = compter_pixels_et_calculer_superficie(image_simulee)

        # Vérifier si les résultats sont corrects
        self.assertEqual(total_pixels, 100 * 100)
        self.assertEqual(white_pixels, 100 * 100)
        self.assertEqual(superficie_km2, 400_000_000)

    def test_supprimer_extension(self):
        # Cas où l'extension est présente
        nom_fichier_avec_extension = "fichier.txt"
        resultat = supprimer_extension(nom_fichier_avec_extension)
        self.assertEqual(resultat, "fichier")

        # Cas où l'extension est absente
        nom_fichier_sans_extension = "fichier"
        resultat = supprimer_extension(nom_fichier_sans_extension)
        self.assertEqual(resultat, "fichier")

    @patch("builtins.print")  # Pour capturer les appels à la fonction print
    def test_main(self, mock_print):
        # Créer un fichier Excel de test
        from openpyxl import Workbook
        test_workbook_path = "test_workbook.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = "Image"
        sheet['A2'] = "fichier.txt"
        workbook.save(test_workbook_path)

        # Appeler la fonction main avec le fichier de test
        main(test_workbook_path)

        # Vérifier si la cellule A1 a été modifiée
        modified_workbook = Workbook()
        modified_workbook = openpyxl.load_workbook(test_workbook_path)
        modified_sheet = modified_workbook.active
        modified_value = modified_sheet['A1'].value
        self.assertEqual(modified_value, "Image")

        # Vérifier si la cellule A2 a été modifiée
        modified_value_a2 = modified_sheet['A2'].value
        self.assertEqual(modified_value_a2, "fichier")

        # Vérifier si la fonction print a été appelée
        self.assertTrue(mock_print.called)

    def setUp(self):
        # Créer des objets factices pour simuler l'interface utilisateur
        self.liste_essences = MagicMock()
        self.resultat_text = MagicMock()

    def test_afficher_resultats(self):
        # Définir les conditions initiales pour la simulation
        essence_selectionnee = "EssenceTest"
        self.liste_essences.get.return_value = essence_selectionnee

        # Appeler la fonction à tester
        afficher_resultats(self.liste_essences, self.resultat_text)

        # Vérifier si les méthodes de l'objet resultat_text ont été appelées correctement
        
        self.resultat_text.delete.assert_called_once_with(1.0, "end")
        self.resultat_text.insert.assert_called()

        # Vérifier si la méthode get de liste_essences a été appelée
        self.liste_essences.get.assert_called_once()

    def test_modifier_chemins_dans_fichier(self):
            # Créez un dossier temporaire
            tmpdir = tempfile.mkdtemp()

            # Créez des chemins absolus fictifs
            img1_path = os.path.join(tmpdir, 'image1.jpg')
            img2_path = os.path.join(tmpdir, 'image2.png')
            img3_path = os.path.join(tmpdir, 'subfolder', 'image3.gif')
            os.makedirs(os.path.join(tmpdir, 'subfolder'))
            open(img1_path, 'w').close()
            open(img2_path, 'w').close()
            open(img3_path, 'w').close()

            # Chemins absolus d'entrée et de sortie
            fichier_entree = os.path.join(tmpdir, 'input.txt')
            fichier_sortie = os.path.join(tmpdir, 'output.txt')

            # Écrivez les chemins absolus dans le fichier d'entrée
            with open(fichier_entree, 'w') as f:
                f.write('\n'.join([img1_path, img2_path, img3_path]))

            # Exécutez la fonction à tester
            dossier_base = tmpdir
            modifier_chemins_dans_fichier(fichier_entree, fichier_sortie, dossier_base)

            # Lisez les chemins relatifs à partir du fichier de sortie
            with open(fichier_sortie, 'r') as f:
                chemins_relatifs = f.read().splitlines()

            # Vérifiez que les chemins relatifs sont corrects
            self.assertEqual(chemins_relatifs, ['image1.jpg', 'image2.png', 'subfolder/image3.gif'])

            # Nettoyez le dossier temporaire
            shutil.rmtree(tmpdir)


if __name__ == '__main__':
    unittest.main()

####################################################################################################################


